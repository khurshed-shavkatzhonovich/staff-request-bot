from django.contrib import admin
from .models import StaffRequest
from django.utils.html import format_html
from django.urls import reverse
from django.http import HttpResponse
import csv
import openpyxl
from openpyxl.utils import get_column_letter

class ExportAsCSV:
    def export_as_csv(self, request, queryset):
        meta = self.model._meta
        field_names = [field.name for field in meta.fields]

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename={meta.model_name}.csv'
        writer = csv.writer(response)
        writer.writerow(field_names)
        for obj in queryset:
            writer.writerow([getattr(obj, field) for field in field_names])
        return response

    export_as_csv.short_description = "Экспорт в CSV"

    def export_as_excel(self, request, queryset):
        meta = self.model._meta
        fields = [(field.name, field.verbose_name) for field in meta.fields]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Данные"

        # Заголовки
        for col_num, (_, verbose_name) in enumerate(fields, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = verbose_name

        # Данные
        for row_num, obj in enumerate(queryset, 2):
            for col_num, (field_name, _) in enumerate(fields, 1):
                value = getattr(obj, field_name)
                ws.cell(row=row_num, column=col_num, value=str(value))

        # Ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={meta.model_name}.xlsx'
        wb.save(response)
        return response

    export_as_excel.short_description = "Экспорт в Excel"

@admin.register(StaffRequest)
class StaffRequestAdmin(admin.ModelAdmin, ExportAsCSV):
    actions = ['export_as_csv', 'export_as_excel']

    list_display = (
        'id', 
        'get_request_type_display', 
        'telegram_user_info',
        'site',
        'description',
        'amount',
        'equipment',
        'created_at',
        'get_status_display',
        'view_link'
    )
    list_filter = (
        'request_type', 
        'status',
        ('created_at', admin.DateFieldListFilter),
    )
    search_fields = (
        'telegram_user_id',
        'telegram_username',
        'full_name',
        'site',
        'equipment',
    )
    list_per_page = 20
    readonly_fields = ('created_at', 'updated_at')

    
    fieldsets = (
        ('Основная информация', {
            'fields': ('request_type', 'status', 'telegram_user_id', 'telegram_username', 'full_name')
        }),
        ('Детали заявки', {
            'fields': ('site', 'equipment', 'description', 'amount')
        }),
        ('Модерация', {
            'fields': ('approved_by', 'approved_at', 'authorized_by', 'authorized_at'),
            'classes': ('collapse',)
        }),
        ('Технические данные', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def telegram_user_info(self, obj):
        return format_html(
            'ID: {}<br>@{}<br>{}',
            obj.telegram_user_id,
            obj.telegram_username or '-',
            obj.full_name or '-'
        )
    telegram_user_info.short_description = 'Пользователь'

    def view_link(self, obj):
        return format_html(
            '<a class="button" href="{}">Просмотр</a>',
            reverse('admin:staffapp_staffrequest_change', args=[obj.id])
        )
    view_link.short_description = ''